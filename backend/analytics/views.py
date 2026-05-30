from rest_framework import generics, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Count, Avg, Q, F
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta
from datetime import datetime
from complaints.models import Complaint, ComplaintStatus, Category
from complaints.serializers import ComplaintListSerializer
from .models import SatisfactionSurvey
from .serializers import SatisfactionSurveySerializer
from accounts.models import UserRole
from accounts.roles import filter_complaints_for_user


def _period_range(period: str, year: int, value: int | None):
    """
    period: monthly|quarterly|semiannual|annual
    value: month (1-12) if monthly, quarter (1-4) if quarterly, half (1-2) if semiannual
    """
    period = (period or '').lower().strip()
    if year < 2000 or year > 2100:
        raise ValueError("Année invalide.")

    if period == 'monthly':
        month = int(value or 0)
        if month < 1 or month > 12:
            raise ValueError("Mois invalide (1-12).")
        start = datetime(year, month, 1, tzinfo=timezone.get_current_timezone())
        end = datetime(year + (1 if month == 12 else 0), (1 if month == 12 else month + 1), 1, tzinfo=timezone.get_current_timezone())
        return start, end

    if period == 'quarterly':
        q = int(value or 0)
        if q < 1 or q > 4:
            raise ValueError("Trimestre invalide (1-4).")
        start_month = (q - 1) * 3 + 1
        start = datetime(year, start_month, 1, tzinfo=timezone.get_current_timezone())
        end_month = start_month + 3
        end = datetime(year + (1 if end_month > 12 else 0), (end_month - 12 if end_month > 12 else end_month), 1, tzinfo=timezone.get_current_timezone())
        return start, end

    if period == 'semiannual':
        half = int(value or 0)
        if half not in (1, 2):
            raise ValueError("Semestre invalide (1-2).")
        start_month = 1 if half == 1 else 7
        start = datetime(year, start_month, 1, tzinfo=timezone.get_current_timezone())
        end = datetime(year, 7, 1, tzinfo=timezone.get_current_timezone()) if half == 1 else datetime(year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
        return start, end

    if period == 'annual':
        start = datetime(year, 1, 1, tzinfo=timezone.get_current_timezone())
        end = datetime(year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
        return start, end

    raise ValueError("Période invalide (monthly|quarterly|semiannual|annual).")


class DashboardView(APIView):
    """Tableau de bord avec KPIs"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        qs = filter_complaints_for_user(user, Complaint.objects.all())

        now = timezone.now()
        thirty_days_ago = now - timedelta(days=30)

        open_statuses = [
            ComplaintStatus.SOUMISE,
            ComplaintStatus.ACCUSEE,
            ComplaintStatus.INSTRUITE,
            ComplaintStatus.AFFECTEE,
            ComplaintStatus.EN_TRAITEMENT,
            ComplaintStatus.ESCALADEE,
            ComplaintStatus.ARBITREE,
        ]

        total = qs.count()
        open_count = qs.filter(status__in=open_statuses).count()
        resolved = qs.filter(status=ComplaintStatus.RESOLUE).count()
        overdue = qs.filter(is_overdue=True).count()

        # Average resolution time (in hours)
        resolved_complaints = qs.filter(resolved_at__isnull=False)
        if resolved_complaints.exists():
            total_hours = sum(
                (c.resolved_at - c.created_at).total_seconds() / 3600
                for c in resolved_complaints[:100]
            )
            avg_resolution = round(total_hours / min(resolved_complaints.count(), 100), 1)
        else:
            avg_resolution = 0

        # Satisfaction average
        satisfaction_avg = SatisfactionSurvey.objects.aggregate(avg=Avg('rating'))['avg'] or 0

        # By status
        by_status = dict(qs.values_list('status').annotate(count=Count('id')).values_list('status', 'count'))

        # By priority
        by_priority = dict(qs.values_list('priority').annotate(count=Count('id')).values_list('priority', 'count'))

        # By category (top 8)
        by_category = list(
            qs.filter(category__isnull=False)
            .values('category__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )

        # By month (last 12 months)
        by_month = list(
            qs.filter(created_at__gte=now - timedelta(days=365))
            .annotate(month=TruncMonth('created_at'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )
        # Serialize months
        by_month = [{'month': m['month'].strftime('%Y-%m'), 'count': m['count']} for m in by_month]

        # By channel
        by_channel = dict(qs.values_list('channel').annotate(count=Count('id')).values_list('channel', 'count'))

        # By establishment (top 10)
        by_establishment = list(
            qs.filter(establishment__isnull=False)
            .values('establishment__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )

        # Recent complaints
        recent = ComplaintListSerializer(qs[:5], many=True).data

        return Response({
            'total_complaints': total,
            'open_complaints': open_count,
            'resolved_complaints': resolved,
            'overdue_complaints': overdue,
            'avg_resolution_time': avg_resolution,
            'satisfaction_avg': round(satisfaction_avg, 1),
            'complaints_by_status': by_status,
            'complaints_by_priority': by_priority,
            'complaints_by_category': by_category,
            'complaints_by_month': by_month,
            'complaints_by_channel': by_channel,
            'complaints_by_establishment': by_establishment,
            'recent_complaints': recent,
        })


class SatisfactionCreateView(generics.CreateAPIView):
    """Soumettre une enquête de satisfaction"""
    serializer_class = SatisfactionSurveySerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class SatisfactionListView(generics.ListAPIView):
    """Liste des enquêtes de satisfaction"""
    serializer_class = SatisfactionSurveySerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = SatisfactionSurvey.objects.all()


class PublicStatsView(APIView):
    """Statistiques publiques pour la landing page"""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        total = Complaint.objects.count()
        resolved = Complaint.objects.filter(status__in=[
            ComplaintStatus.RESOLUE,
            ComplaintStatus.CLOTUREE,
        ]).count()
        satisfaction = SatisfactionSurvey.objects.aggregate(avg=Avg('rating'))['avg'] or 0

        return Response({
            'total_complaints': total,
            'resolved_complaints': resolved,
            'resolution_rate': round((resolved / total * 100) if total > 0 else 0, 1),
            'satisfaction_avg': round(satisfaction, 1),
        })


class ExportStatsView(APIView):
    """
    Export statistiques (CABINET / ADMIN).
    Query params:
      - format=pdf|xlsx (default xlsx)
      - period=monthly|quarterly|semiannual|annual (default monthly)
      - year=YYYY (default current year)
      - value=month|quarter|half (selon period)
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role not in [UserRole.CABINET, UserRole.ADMIN_PLATEFORME]:
            return Response({'error': 'Accès réservé au Cabinet ou Admin.'}, status=403)

        fmt = (request.query_params.get('format') or 'xlsx').lower().strip()
        period = (request.query_params.get('period') or 'monthly').lower().strip()
        year = int(request.query_params.get('year') or timezone.now().year)
        value = request.query_params.get('value')
        value_int = int(value) if value is not None and str(value).strip() != '' else None

        try:
            start, end = _period_range(period, year, value_int)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=400)

        qs = Complaint.objects.filter(created_at__gte=start, created_at__lt=end)

        # KPIs
        open_statuses = [
            ComplaintStatus.SOUMISE,
            ComplaintStatus.ACCUSEE,
            ComplaintStatus.INSTRUITE,
            ComplaintStatus.AFFECTEE,
            ComplaintStatus.EN_TRAITEMENT,
            ComplaintStatus.ESCALADEE,
            ComplaintStatus.ARBITREE,
        ]
        total = qs.count()
        open_count = qs.filter(status__in=open_statuses).count()
        resolved = qs.filter(status=ComplaintStatus.RESOLUE).count()
        overdue = qs.filter(is_overdue=True).count()

        by_status = dict(qs.values_list('status').annotate(count=Count('id')).values_list('status', 'count'))
        by_priority = dict(qs.values_list('priority').annotate(count=Count('id')).values_list('priority', 'count'))
        by_channel = dict(qs.values_list('channel').annotate(count=Count('id')).values_list('channel', 'count'))
        by_category = list(
            qs.filter(category__isnull=False)
            .values('category__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        title = f"PGP-USS — Export statistiques ({period}) {year}"
        if period in ('monthly', 'quarterly', 'semiannual') and value_int:
            title += f" — {value_int}"

        if fmt == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Synthèse"

            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:D1")

            ws["A3"] = "Période"
            ws["B3"] = f"{start.date()} → {end.date()}"
            ws["A4"] = "Total plaintes"
            ws["B4"] = total
            ws["A5"] = "En cours"
            ws["B5"] = open_count
            ws["A6"] = "Résolues"
            ws["B6"] = resolved
            ws["A7"] = "En retard"
            ws["B7"] = overdue

            for r in range(3, 8):
                ws[f"A{r}"].font = Font(bold=True)
                ws[f"A{r}"].alignment = Alignment(horizontal="left")

            # Breakdowns
            ws2 = wb.create_sheet("Par statut")
            ws2.append(["Statut", "Nombre"])
            for k, v in sorted(by_status.items(), key=lambda x: x[1], reverse=True):
                ws2.append([k, v])

            ws3 = wb.create_sheet("Par priorité")
            ws3.append(["Priorité", "Nombre"])
            for k, v in sorted(by_priority.items(), key=lambda x: x[0]):
                ws3.append([k, v])

            ws4 = wb.create_sheet("Par canal")
            ws4.append(["Canal", "Nombre"])
            for k, v in sorted(by_channel.items(), key=lambda x: x[1], reverse=True):
                ws4.append([k, v])

            ws5 = wb.create_sheet("Par catégorie")
            ws5.append(["Catégorie", "Nombre"])
            for row in by_category:
                ws5.append([row.get("category__name"), row.get("count")])

            out = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"pgpuss_stats_{period}_{year}_{value_int or ''}.xlsx".replace("__", "_").strip("_")
            out["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(out)
            return out

        if fmt == 'pdf':
            import io
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas

            buf = io.BytesIO()
            c = canvas.Canvas(buf, pagesize=A4)
            width, height = A4
            y = height - 60
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40, y, title)
            y -= 22
            c.setFont("Helvetica", 10)
            c.drawString(40, y, f"Période: {start.date()} → {end.date()}")
            y -= 18

            def line(label, val):
                nonlocal y
                c.setFont("Helvetica-Bold", 10)
                c.drawString(40, y, f"{label}:")
                c.setFont("Helvetica", 10)
                c.drawString(180, y, str(val))
                y -= 14

            line("Total plaintes", total)
            line("En cours", open_count)
            line("Résolues", resolved)
            line("En retard", overdue)
            y -= 10

            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, "Répartition par statut (top)")
            y -= 16
            c.setFont("Helvetica", 9)
            for k, v in sorted(by_status.items(), key=lambda x: x[1], reverse=True)[:12]:
                c.drawString(50, y, f"- {k}: {v}")
                y -= 12
                if y < 80:
                    c.showPage()
                    y = height - 60

            c.showPage()
            c.save()
            pdf = buf.getvalue()
            buf.close()
            resp = HttpResponse(content_type="application/pdf")
            filename = f"pgpuss_stats_{period}_{year}_{value_int or ''}.pdf".replace("__", "_").strip("_")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            resp.write(pdf)
            return resp

        return Response({'error': 'format invalide (pdf|xlsx).'}, status=400)
